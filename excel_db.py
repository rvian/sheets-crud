from datetime import datetime
from pytz import timezone
import openpyxl as op

fuso_horario = timezone('America/Bahia')
now = datetime.now(tz= fuso_horario).strftime("%d/%m/%Y %H:%M:%S")



# headerManual = ('ID', 'NOME_EMPRESARIAL', 'CNPJ', 'CAPITAL_SOCIAL', 'ENDEREÇO', 'CEP', 'CIDADE', 'UF', 'EMAIL', 'TELEFONE', 'BANCO', 'AGÊNCIA', 'N_CONTA', 'REPRESENTANTE_LEGAL', 'CPF_REPRESENTANTE',
#                 'ENDEREÇO_REPRESENTANTE', 'CEP_REPRESENTANTE', 'CIDADE_REPRESENTANTE', 'UF_REPRESENTANTE', 'EMAIL_REPRESENTANTE', 'FONE_REPRESENTANTE', 'DATA_SOLICITACAO', 'DATA_RECEBIMENTO', 'DATA_ASSINATURA')
# value = {  # "ID": varId,
#     "ID": 2,
#     "NOME_EMPRESARIAL": "EDITAAAADO",
#     "CNPJ": "71.s.527/sss-35",
#     "CAPITAL_SOCIAL": "ss1095829378.49",
#     "ENDEREÇO": "RUssssA BOA VISTA, 280, CENTRO, PAVMTO8 E 9",
#     "CEP": "01014908",
#     "CIDADE": "SÃO PAULO",
#     "UF": "SP",
#     "EMAIL": "FISCAL@TENDA.COM",
#     "TELEFONE": "1131112536",
#     "BANCO": "BANCO DO BRASIL",
#     "AGÊNCIA": "033",
#     "N_CONTA": "5794331",
#     "REPRESENTANTE_LEGAL": "VINICIUS FARAJ",
#     "CPF_REPRESENTANTE": "***378228**",
#     "ENDEREÇO_REPRESENTANTE": "ND",
#     "CEP_REPRESENTANTE": "ND",
#     "CIDADE_REPRESENTANTE": "ND",
#     "UF_REPRESENTANTE": "ND",
#     "EMAIL_REPRESENTANTE": "ND",
#     "FONE_REPRESENTANTE": "ND",
#     "DATA_SOLICITACAO": "26/02/2022",
#     "DATA_RECEBIMENTO": "29/12/2021",
#     "DATA_ASSINATURA": "31/12/2021",
#     "DATA_DOCUMENTO": "23/12/2021",
#     "DATA_CRIACAO": now.strftime("%d/%m/%Y")
# }

wb = op.load_workbook('dados.xlsx')
ws = wb["base"]


def listar_empresas():
    return list(ws.values)

def buscar_empresa( index ):
    data = listar_empresas()[1:]
    dict = ListToDict( data[index] )
    dict["index"] = index
    return dict


# Converte uma lista (linha selecionada) para dict
def ListToDict(row):
    # TODO: pegar keys automaticamente
    dictKey = [c.value for c in next(wb["base"].iter_rows(min_row=1, max_row=1))]
    dictValue = row

    return dict(zip(dictKey, dictValue))


# novaEmpresa dict
def criar_empresa(novaEmpresa):
    input = list(novaEmpresa.values())
    input[0] = ws.max_row + 1 #set id
    input[25] = now
    ws.append(input)
    wb.save('dados.xlsx')

def excluir_empresa(index):
    if index < 1:
        return
    ws.delete_rows(index, 1)
    wb.save('dados.xlsx')


def editar_empresa(index,  novaEmpresa):
    inputList = list(novaEmpresa.values())
    inputList[24] = now
    for col in range(0, ws.max_column):
        ws.cell(row = index + 2, column=col + 1).value = inputList[col]
    wb.save('dados.xlsx')
    
    
def limpar_none():
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row, col).value == None:
                ws.cell(row, col).value = "ㅤ" # define char invisivel para celulas vazias
    wb.save('dados.xlsx')

limpar_none()